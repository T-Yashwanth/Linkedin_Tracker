import argparse
import json
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font
from dateutil import parser as dateparser

from src.gmail_client import get_gmail_service
from src.parser import get_html_body, parse_application_email
from src.dice_parser import parse_dice_subject
from src.sent_matcher import fetch_sent_index, find_hiring_managers, find_reachout_contacts
from src.phone_lookup import find_phone_for_email, get_own_phone_numbers

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PROCESSED_FILE = os.path.join(BASE_DIR, 'data', 'processed_ids.json')
DEFAULT_TRACKER = os.path.join(BASE_DIR, 'data', 'Job_Tracker.xlsx')


def parse_linkedin_message(msg, headers):
    subject = headers.get('Subject', '')
    html = get_html_body(msg['payload'])
    data = parse_application_email(subject, html)
    return {
        'company': data['company'],
        'title': data['title'],
        'applied_date': data['applied_date'],
        'job_link': data['job_link'],
    }


def parse_dice_message(msg, headers):
    subject = headers.get('Subject', '')
    data = parse_dice_subject(subject)
    return {
        'company': data['company'],
        'title': data['title'],
        'applied_date': None,
        'job_link': None,
    }


SOURCES = [
    {
        'label': 'LinkedIn',
        'query': 'from:jobs-noreply@linkedin.com "your application was sent to"',
        'parse': parse_linkedin_message,
        'source_value': 'LinkedIn',
    },
    {
        'label': 'Dice',
        'query': 'from:applyonline@dice.com "Application for"',
        'parse': parse_dice_message,
        'source_value': 'Dice',
    },
]
SOURCES_BY_LABEL = {s['label'].lower(): s for s in SOURCES}
REACHOUT_PLATFORM_VALUE = 'Email Reach Out'

HEADERS = [
    'S.no', 'Date', 'Title', 'Company Full Name', 'Platform',
    'Website Applied', 'Hiring Manager In Linkedin', 'Company Email',
    'Contact Number For Job Post', 'Comment Section',
]

# Older trackers stored 'Yes' in the platform column back when LinkedIn was
# the only source; normalize that to the source label on every run.
LEGACY_PLATFORM_VALUES = {'yes': 'LinkedIn'}
COLUMN_WIDTHS = {'D': 15.5, 'E': 14.3, 'F': 17.2, 'G': 15.8, 'H': 13.3, 'I': 22.4, 'J': 17.8}


def load_processed():
    if os.path.exists(PROCESSED_FILE):
        with open(PROCESSED_FILE) as f:
            return set(json.load(f))
    return set()


def save_processed(ids):
    os.makedirs(os.path.dirname(PROCESSED_FILE), exist_ok=True)
    with open(PROCESSED_FILE, 'w') as f:
        json.dump(sorted(ids), f, indent=2)


def create_blank_tracker(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)


def ensure_tracker(path, rebuild):
    if rebuild and os.path.exists(path):
        os.remove(path)
    if not os.path.exists(path):
        create_blank_tracker(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    for col, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)
    return wb


def last_data_row(ws):
    """Last row (2-indexed) that has any data in columns B-J, scanning from
    the sheet's max_row down. Ignores column A (S.no) since manually-added
    rows may not have it filled in."""
    for r in range(ws.max_row, 1, -1):
        if any(ws.cell(row=r, column=c).value not in (None, '') for c in range(2, 11)):
            return r
    return 1


def read_existing_rows(ws):
    """Read existing data rows (columns B-J) into a list of {'date',
    'values'}, migrating legacy platform values along the way. Returns
    (existing_rows, last_row)."""
    existing_rows = []
    last_row = last_data_row(ws)
    for r in range(2, last_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(2, 11)]
        platform = values[3]
        if isinstance(platform, str) and platform.strip().lower() in LEGACY_PLATFORM_VALUES:
            values[3] = LEGACY_PLATFORM_VALUES[platform.strip().lower()]
        date_str = values[0]
        try:
            row_date = dateparser.parse(date_str).date() if date_str else datetime.max.date()
        except (ValueError, OverflowError):
            row_date = datetime.max.date()
        existing_rows.append({'date': row_date, 'values': values})
    return existing_rows, last_row


def fetch_all_messages(service, query, max_results):
    messages = []
    request = service.users().messages().list(userId='me', q=query, maxResults=min(max_results, 500))
    while request is not None and len(messages) < max_results:
        response = request.execute()
        messages.extend(response.get('messages', []))
        request = service.users().messages().list_next(previous_request=request, previous_response=response)
    return messages[:max_results]


def main():
    ap = argparse.ArgumentParser(description='Fetch LinkedIn application emails and update the tracker.')
    ap.add_argument('--tracker', default=DEFAULT_TRACKER, help='Path to the tracker xlsx to update')
    ap.add_argument('--max-results', type=int, default=2000, help='Max Gmail messages to scan')
    ap.add_argument('--since', default=None, help='Only include applications on/after this date, e.g. 2026-06-24')
    ap.add_argument('--rebuild', action='store_true', help='Start the tracker fresh instead of appending')
    ap.add_argument('--no-hiring-manager', action='store_true', help='Skip matching hiring manager name/email from Sent mail')
    ap.add_argument('--dry-run', action='store_true', help='Parse and print results without writing to xlsx')
    ap.add_argument('--source', choices=['all', 'reachout'] + list(SOURCES_BY_LABEL), default='all',
                     help='Only fetch new applications from this source (default: all)')
    ap.add_argument('--include-reachout', action='store_true',
                     help='Also log recruiters/contacts you emailed on a company domain that are not tied to an existing application')
    ap.add_argument('--include-phone', action='store_true',
                     help='Look up recruiter phone numbers from Inbox replies for hiring-manager and reach-out contacts, and backfill existing blank entries (slow: one extra Gmail search per unique contact email)')
    args = ap.parse_args()

    since_date = dateparser.parse(args.since).date() if args.since else None
    if args.source == 'reachout':
        active_sources = []
        include_reachout = True
    else:
        active_sources = SOURCES if args.source == 'all' else [SOURCES_BY_LABEL[args.source]]
        include_reachout = args.include_reachout

    service = get_gmail_service()
    processed = set() if args.rebuild else load_processed()

    existing_rows = []
    if args.dry_run:
        # Read-only preview: mirror what a real run would see, without
        # creating or modifying the tracker file. A --rebuild run ignores
        # existing data entirely, so the preview does too.
        wb = None
        ws = None
        if not args.rebuild and os.path.exists(args.tracker):
            existing_rows, _ = read_existing_rows(openpyxl.load_workbook(args.tracker).active)
    else:
        wb = ensure_tracker(args.tracker, args.rebuild)
        ws = wb.active
        existing_rows, last_row = read_existing_rows(ws)
        if last_row >= 2:
            ws.delete_rows(2, last_row - 1)

    sent_index = []
    need_sent_index = (not args.no_hiring_manager) or include_reachout
    if need_sent_index:
        idx_since = since_date or datetime(2020, 1, 1).date()
        print('Indexing your Sent mail...')
        sent_index = fetch_sent_index(service, fetch_all_messages, idx_since)
        print(f'Indexed {len(sent_index)} sent message(s).')

    known_emails = set()
    for r in existing_rows:
        emails_cell = r['values'][6]
        if emails_cell:
            known_emails.update(e.strip() for e in str(emails_cell).split(';') if e.strip())

    phone_cache = {}
    own_numbers = set()

    def lookup_phone(email):
        key = email.lower().strip()
        if key not in phone_cache:
            phone_cache[key] = find_phone_for_email(
                service, fetch_all_messages, key, exclude_digits=own_numbers)
        return phone_cache[key]

    if args.include_phone:
        own_numbers = get_own_phone_numbers(service, fetch_all_messages)
        if own_numbers:
            print(f'Detected {len(own_numbers)} of your own phone number(s) from your Sent-mail signature; they will never be recorded as a recruiter number.')
        print('Looking up recruiter phone numbers from Inbox replies (this can take a while)...')
        for row in existing_rows:
            values = row['values']
            if values[7]:
                continue
            for email in (values[6] or '').split(';'):
                email = email.strip()
                if not email:
                    continue
                phone = lookup_phone(email)
                if phone:
                    values[7] = phone
                    break

    new_rows = []
    new_count = 0
    skipped_unparsed = 0

    for source in active_sources:
        query = source['query']
        if since_date:
            query += f' after:{since_date.strftime("%Y/%m/%d")}'

        messages = fetch_all_messages(service, query, args.max_results)
        print(f"Found {len(messages)} matching {source['label']} email(s) in Gmail.")

        for m in messages:
            msg_id = m['id']
            if msg_id in processed:
                continue

            msg = service.users().messages().get(userId='me', id=msg_id, format='full').execute()
            headers = {h['name']: h['value'] for h in msg['payload'].get('headers', [])}
            data = source['parse'](msg, headers)

            if not data['company']:
                skipped_unparsed += 1
                processed.add(msg_id)
                continue

            if data['applied_date']:
                try:
                    applied_dt = dateparser.parse(data['applied_date'])
                except (ValueError, OverflowError):
                    applied_dt = datetime.fromtimestamp(int(msg['internalDate']) / 1000)
            else:
                applied_dt = datetime.fromtimestamp(int(msg['internalDate']) / 1000)

            if since_date and applied_dt.date() < since_date:
                processed.add(msg_id)
                continue

            hiring_managers = []
            if not args.no_hiring_manager:
                hiring_managers = find_hiring_managers(data['company'], applied_dt.date(), sent_index)
                known_emails.update(hm['email'] for hm in hiring_managers)

            hm_names = '; '.join(hm['name'] for hm in hiring_managers)
            hm_emails = '; '.join(hm['email'] for hm in hiring_managers)

            hm_phone = ''
            if args.include_phone:
                for hm in hiring_managers:
                    phone = lookup_phone(hm['email'])
                    if phone:
                        hm_phone = phone
                        break

            if args.dry_run:
                hm = f" | HM: {hm_names} <{hm_emails}>" if hiring_managers else ''
                phone_note = f" | Phone: {hm_phone}" if hm_phone else ''
                print(source['label'], '|', applied_dt.strftime('%m/%d/%Y'), '|', data['company'], '|', data['title'], '|', data['job_link'], hm, phone_note)
            else:
                values = [
                    applied_dt.strftime('%m/%d/%Y'),
                    data['title'] or '',
                    data['company'] or '',
                    source['source_value'],
                    data['job_link'] or '',
                    hm_names,
                    hm_emails,
                    hm_phone or None,
                    None,
                ]
                new_rows.append({'date': applied_dt.date(), 'values': values})

            new_count += 1
            processed.add(msg_id)

    reachout_count = 0
    if include_reachout:
        contacts = find_reachout_contacts(sent_index, known_emails)
        if since_date:
            contacts = [c for c in contacts if c['date'] >= since_date]
        for contact in contacts:
            contact_phone = lookup_phone(contact['email']) if args.include_phone else None

            if args.dry_run:
                phone_note = f" | Phone: {contact_phone}" if contact_phone else ''
                print('Email Reach Out', '|', contact['date'].strftime('%m/%d/%Y'), '|', contact['company_guess'], '|', contact['name'], '|', contact['email'], phone_note)
            else:
                values = [
                    contact['date'].strftime('%m/%d/%Y'),
                    '',
                    contact['company_guess'],
                    REACHOUT_PLATFORM_VALUE,
                    '',
                    contact['name'],
                    contact['email'],
                    contact_phone,
                    None,
                ]
                new_rows.append({'date': contact['date'], 'values': values})
            reachout_count += 1
        print(f'Found {reachout_count} reach-out contact(s) not already tied to an existing application.')

    if not args.dry_run:
        all_rows = existing_rows + new_rows
        all_rows.sort(key=lambda r: r['date'])
        for i, r in enumerate(all_rows, start=1):
            ws.cell(row=i + 1, column=1, value=i)
            for c, val in enumerate(r['values'], start=2):
                ws.cell(row=i + 1, column=c, value=val)

        wb.save(args.tracker)
        save_processed(processed)
        print(f'Added {new_count} new application(s) + {reachout_count} reach-out contact(s); tracker now has {len(all_rows)} row(s), sorted by date, saved to {args.tracker}')
    else:
        print(f'{new_count} new application(s) + {reachout_count} reach-out contact(s) would be added (dry run, nothing saved).')

    if skipped_unparsed:
        print(f'Skipped {skipped_unparsed} email(s) that could not be parsed (marked as processed).')

    if args.include_phone:
        attempted = len(phone_cache)
        found = sum(1 for v in phone_cache.values() if v)
        print(f'Looked up phone numbers for {attempted} unique contact(s), found {found}.')


if __name__ == '__main__':
    main()
