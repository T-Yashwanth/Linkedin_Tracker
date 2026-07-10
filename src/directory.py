import os
import re
from email.utils import parseaddr

import openpyxl
from openpyxl.styles import Font

from src.sent_matcher import GENERIC_DOMAINS, guess_company_from_domain, name_from_email_local

DIRECTORY_HEADERS = [
    'S.no', 'Domain', 'Company Name', 'Job Title', 'Recruiter Name',
    'Phone', 'Email', 'LinkedIn Profile', 'Positions',
]
DIRECTORY_COLUMN_WIDTHS = {
    'B': 22, 'C': 20, 'D': 26, 'E': 20, 'F': 16, 'G': 30, 'H': 34, 'I': 45,
}

JOB_BOARD_DOMAINS = {
    'dice.com', 'indeed.com', 'ziprecruiter.com', 'glassdoor.com',
    'lever.co', 'greenhouse.io', 'icims.com', 'myworkday.com',
    'smartrecruiters.com',
}

BOT_LOCAL_PART_RE = re.compile(
    r'no-?reply|do-?not-?reply|notifications?|alerts?|mailer-daemon|'
    r'bounce|^auto|jobs-noreply|applyonline|newsletter|marketing|'
    r'updates|digest|calendar-notification|postmaster|support|info|'
    r'^account$|verification|productivity-report|billing|feedback',
    re.IGNORECASE,
)

SUBJECT_PREFIX_RE = re.compile(r'^\s*(re|fwd?)\s*:\s*', re.IGNORECASE)


def clean_subject(subject):
    """Strip repeated Re:/Fwd: prefixes and surrounding whitespace."""
    if not subject:
        return None
    prev = None
    s = subject.strip()
    while prev != s:
        prev = s
        s = SUBJECT_PREFIX_RE.sub('', s).strip()
    return s or None


def is_bot_headers(headers):
    """True if the message carries header-level signals of an automated
    sender (mailing list / bulk mail / auto-generated)."""
    if headers.get('List-Unsubscribe'):
        return True
    precedence = (headers.get('Precedence') or '').lower()
    if precedence in ('bulk', 'junk', 'list'):
        return True
    if headers.get('Auto-Submitted', '').lower() not in ('', 'no'):
        return True
    return False


def is_eligible_contact(email, domain):
    """True if this address looks like a real, domain-specific human
    contact worth tracking (not a personal-mail provider, job board, or
    automated/bot sender)."""
    if not email or '@' not in email:
        return False
    if domain in GENERIC_DOMAINS or domain in JOB_BOARD_DOMAINS:
        return False
    local_part = email.split('@')[0]
    if BOT_LOCAL_PART_RE.search(local_part):
        return False
    return True


def parse_recipients(header_value):
    """Yield (name, email) for every address in a To/Cc header value."""
    if not header_value:
        return
    for chunk in header_value.split(','):
        name, addr = parseaddr(chunk.strip())
        if addr and '@' in addr:
            yield name.strip(), addr.lower().strip()


class ContactBook:
    """Accumulates directory entries across many messages before they're
    merged into the sheet. One entry per unique (lowercase) email."""

    def __init__(self):
        self._contacts = {}

    def add(self, email, name=None, subject=None):
        domain = email.split('@')[-1]
        if not is_eligible_contact(email, domain):
            return

        entry = self._contacts.setdefault(email, {
            'domain': domain,
            'name': None,
            'subjects': set(),
        })

        if name and (not entry['name'] or len(name) > len(entry['name'])):
            entry['name'] = name

        cleaned = clean_subject(subject)
        if cleaned:
            entry['subjects'].add(cleaned)

    def items(self):
        for email, entry in self._contacts.items():
            name = entry['name'] or name_from_email_local(email.split('@')[0]) or ''
            yield email, {
                'domain': entry['domain'],
                'company': guess_company_from_domain(entry['domain']),
                'name': name,
                'subjects': entry['subjects'],
            }


def create_blank_directory(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(DIRECTORY_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col, width in DIRECTORY_COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)


def ensure_directory(path, rebuild):
    if rebuild and os.path.exists(path):
        os.remove(path)
    if not os.path.exists(path):
        create_blank_directory(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    for col, header in enumerate(DIRECTORY_HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)
    return wb


def _last_data_row(ws):
    for r in range(ws.max_row, 1, -1):
        if any(ws.cell(row=r, column=c).value not in (None, '') for c in range(2, 10)):
            return r
    return 1


def read_directory_rows(ws):
    """Read existing directory rows (columns B-I) into {email: values} for
    rows that have an email, plus a separate list of rows with no email
    (kept as-is, e.g. manually added entries missing that field)."""
    by_email = {}
    no_email_rows = []
    last_row = _last_data_row(ws)
    for r in range(2, last_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(2, 10)]
        email = (values[5] or '').strip().lower()  # column G, index 5 within B-I
        if email:
            by_email[email] = values
        else:
            no_email_rows.append(values)
    return by_email, no_email_rows


def merge_contact(existing_values, domain, company, title, name, phone, email,
                   linkedin, subjects):
    """Return updated column values for one contact: blanks get filled,
    non-blank existing values are preserved, Positions is a union-append.
    existing_values may be None for a brand-new contact."""
    if existing_values is None:
        existing_values = [None] * 8

    values = list(existing_values)
    fill = {0: domain, 1: company, 2: title, 3: name, 4: phone, 5: email, 6: linkedin}
    for idx, new_val in fill.items():
        if not values[idx] and new_val:
            values[idx] = new_val

    existing_positions = [p.strip() for p in (values[7] or '').split(';') if p.strip()]
    seen = set(p.lower() for p in existing_positions)
    for pos in sorted(subjects or []):
        if pos.lower() not in seen:
            existing_positions.append(pos)
            seen.add(pos.lower())
    values[7] = '; '.join(existing_positions) if existing_positions else None

    return values


def write_directory(path, by_email, no_email_rows):
    """Rewrite the directory: rows with an email sorted by (domain, email),
    then any manually-added rows lacking an email, preserved as-is."""
    wb = ensure_directory(path, rebuild=False)
    ws = wb.active
    last_row = _last_data_row(ws)
    if last_row >= 2:
        ws.delete_rows(2, last_row - 1)

    ordered = sorted(by_email.items(), key=lambda kv: (kv[1][0] or '', kv[0]))

    row_idx = 2
    sno = 1
    for email, values in ordered:
        ws.cell(row=row_idx, column=1, value=sno)
        for c, val in enumerate(values, start=2):
            ws.cell(row=row_idx, column=c, value=val)
        row_idx += 1
        sno += 1

    for values in no_email_rows:
        ws.cell(row=row_idx, column=1, value=sno)
        for c, val in enumerate(values, start=2):
            ws.cell(row=row_idx, column=c, value=val)
        row_idx += 1
        sno += 1

    wb.save(path)
